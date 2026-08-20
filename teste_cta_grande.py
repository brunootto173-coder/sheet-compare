from openpyxl import load_workbook


arquivo = "/workspaces/sheet-compare/BASE CTA.xlsx"


print("Abrindo arquivo...")

wb = load_workbook(
    arquivo,
    read_only=True,
    data_only=True,
)

print("Arquivo aberto.")

print("Planilhas:")

for nome in wb.sheetnames:
    print("-", nome)

ws = wb[wb.sheetnames[0]]

print("\nDimensão:")
print(ws.max_row, "linhas")
print(ws.max_column, "colunas")

print("\nPrimeiras 10 linhas:")

for numero, linha in enumerate(
    ws.iter_rows(
        max_row=10,
        values_only=True,
    ),
    start=1,
):

    print(numero, linha)

wb.close()

print("\nFinalizado.")
